#!/usr/bin/env python
"""Parse and sort Excel rows by the `Time` column.

Supported canonical time formats:
- BCE years use a leading minus sign, for example `-305`
- CE dates use dot-separated parts, for example `2001.1` or `2001.1.5`
- BCE dates also use dot-separated parts, for example `-305.1.4`
- Time ranges use `", "` as the separator, for example `-221, -206`

The parser also tolerates a few legacy variants found in history notes, such as
`公元前209年7月`, `191110月10日`, and `170万`.
"""

from __future__ import annotations

import argparse
import dataclasses
import math
import re
import shutil
import sys
from copy import copy
from pathlib import Path
from typing import Iterable, Sequence

from openpyxl import load_workbook


TIME_HEADER = "Time"
RANGE_SEPARATOR = ", "

SEASON_MONTH = {
    "春": 3,
    "夏": 6,
    "秋": 9,
    "冬": 12,
}

QUALIFIER_ORDER = {
    "以前": -2,
    "前后": -1,
    "": 0,
    "初": 1,
    "春": 2,
    "夏": 3,
    "秋": 4,
    "冬": 5,
    "末": 6,
    "以后": 7,
}

YEAR_MULTIPLIER = {
    "": 1,
    "千": 1_000,
    "万": 10_000,
    "亿": 100_000_000,
}


@dataclasses.dataclass(frozen=True)
class TimePoint:
    year: float
    month: int | None = None
    day: int | None = None
    qualifier: str = ""
    original: str = ""

    @property
    def sort_key(self) -> tuple[float, int, int, int]:
        month = self.month if self.month is not None else inferred_month(self.qualifier)
        day = self.day if self.day is not None else 0
        qualifier_rank = QUALIFIER_ORDER.get(self.qualifier, 0)
        return (self.year, month, day, qualifier_rank)

    def normalized(self) -> str:
        year_text = format_year(self.year)
        if self.month is None:
            return f"{year_text}{self.qualifier}"
        if self.day is None:
            return f"{year_text}.{self.month}{self.qualifier}"
        return f"{year_text}.{self.month}.{self.day}{self.qualifier}"


@dataclasses.dataclass(frozen=True)
class TimeSpan:
    raw: str
    points: tuple[TimePoint, ...]

    @property
    def normalized(self) -> str:
        return RANGE_SEPARATOR.join(point.normalized() for point in self.points)

    @property
    def sort_key(self) -> tuple[float, int, int, int, float, int, int, int, int]:
        start = self.points[0].sort_key
        end = self.points[-1].sort_key
        return (
            start[0],
            start[1],
            start[2],
            start[3],
            end[0],
            end[1],
            end[2],
            end[3],
            len(self.points),
        )


@dataclasses.dataclass
class CellSnapshot:
    value: object
    style: object
    hyperlink: object
    comment: object


@dataclasses.dataclass
class RowSnapshot:
    index: int
    cells: list[CellSnapshot]
    time_span: TimeSpan


class TimeParseError(ValueError):
    """Raised when a cell value cannot be parsed as a supported time."""


def inferred_month(qualifier: str) -> int:
    if qualifier in SEASON_MONTH:
        return SEASON_MONTH[qualifier]
    if qualifier == "初":
        return 1
    if qualifier == "末":
        return 12
    return 0


def format_year(year: float) -> str:
    if math.isfinite(year) and float(year).is_integer():
        return str(int(year))
    return f"{year:g}"


def clean_time_text(value: object) -> str:
    text = str(value).strip()
    text = text.replace("，", ",").replace("、", ",")
    text = text.replace("－", "-").replace("—", ",")
    text = text.replace("到", ",")
    text = re.sub(r"\s+", " ", text)
    text = re.sub(r"\s*,\s*", RANGE_SEPARATOR, text)
    return text


def parse_year_token(year_text: str, sign: int = 1) -> float:
    match = re.fullmatch(r"(\d+(?:\.\d+)?)([千万亿]?)", year_text)
    if not match:
        raise TimeParseError(f"Unsupported year token: {year_text!r}")

    number_text, unit = match.groups()
    magnitude = float(number_text) * YEAR_MULTIPLIER[unit]
    magnitude = int(magnitude) if magnitude.is_integer() else magnitude
    return sign * magnitude


def parse_canonical_time(text: str) -> TimePoint | None:
    match = re.fullmatch(
        r"(?P<year>-?\d+)(?:\.(?P<month>\d{1,2})(?:\.(?P<day>\d{1,2}))?)?(?P<qualifier>前后|以前|以后|初|末|春|夏|秋|冬)?",
        text,
    )
    if not match:
        return None

    year = int(match.group("year"))
    month_text = match.group("month")
    day_text = match.group("day")
    qualifier = match.group("qualifier") or ""

    month = int(month_text) if month_text is not None else None
    day = int(day_text) if day_text is not None else None
    return TimePoint(
        year=year,
        month=month,
        day=day,
        qualifier=qualifier,
        original=text,
    )


def parse_legacy_time(text: str) -> TimePoint | None:
    match = re.fullmatch(
        r"(?P<era>公元前|前)?(?P<year>\d+(?:\.\d+)?)(?P<unit>[千万亿]?)年?(?:(?P<month>\d{1,2})月)?(?:(?P<day>\d{1,2})日)?(?P<qualifier>前后|以前|以后|初|末|春|夏|秋|冬)?",
        text,
    )
    if match:
        sign = -1 if match.group("era") in {"公元前", "前"} else 1
        year = parse_year_token(match.group("year") + (match.group("unit") or ""), sign=sign)
        month_text = match.group("month")
        day_text = match.group("day")
        qualifier = match.group("qualifier") or ""

        month = int(month_text) if month_text is not None else None
        day = int(day_text) if day_text is not None else None
        return TimePoint(
            year=year,
            month=month,
            day=day,
            qualifier=qualifier,
            original=text,
        )

    compact = re.fullmatch(
        r"(?P<year>\d{4})(?P<month>\d{1,2})月(?:(?P<day>\d{1,2})日)?(?P<qualifier>前后|以前|以后|初|末|春|夏|秋|冬)?",
        text,
    )
    if compact:
        return TimePoint(
            year=int(compact.group("year")),
            month=int(compact.group("month")),
            day=int(compact.group("day")) if compact.group("day") else None,
            qualifier=compact.group("qualifier") or "",
            original=text,
        )

    bare_year = re.fullmatch(r"(?P<sign>-)?(?P<year>\d+(?:\.\d+)?)(?P<unit>[千万亿])", text)
    if bare_year:
        sign = -1 if bare_year.group("sign") else 1
        year = parse_year_token(bare_year.group("year") + bare_year.group("unit"), sign=sign)
        return TimePoint(year=year, original=text)

    return None


def parse_time_point(text: str) -> TimePoint:
    canonical = parse_canonical_time(text)
    if canonical is not None:
        return canonical

    legacy = parse_legacy_time(text)
    if legacy is not None:
        return legacy

    raise TimeParseError(f"Unsupported time format: {text!r}")


def parse_time_span(value: object) -> TimeSpan:
    if value is None:
        raise TimeParseError("Empty time cell")

    if isinstance(value, (int, float)) and not isinstance(value, bool):
        text = format_year(value)
    else:
        text = clean_time_text(value)

    parts = [part.strip() for part in text.split(",") if part.strip()]
    if not parts:
        raise TimeParseError("Empty time cell")

    points = tuple(parse_time_point(part) for part in parts)
    return TimeSpan(raw=text, points=points)


def build_row_snapshot(ws, row_idx: int, time_col_idx: int, max_col: int) -> RowSnapshot:
    time_value = ws.cell(row_idx, time_col_idx).value
    time_span = parse_time_span(time_value)
    cells = []
    for col_idx in range(1, max_col + 1):
        cell = ws.cell(row_idx, col_idx)
        cells.append(
            CellSnapshot(
                value=cell.value,
                style=copy(cell._style),
                hyperlink=copy(cell.hyperlink) if cell.hyperlink else None,
                comment=copy(cell.comment) if cell.comment else None,
            )
        )
    return RowSnapshot(index=row_idx, cells=cells, time_span=time_span)


def find_time_column(ws, header_row: int, explicit_name: str) -> int:
    for col_idx in range(1, ws.max_column + 1):
        value = ws.cell(header_row, col_idx).value
        if str(value).strip() == explicit_name:
            return col_idx
    raise KeyError(f"Could not find header {explicit_name!r} in row {header_row}")


def sort_sheet_rows(
    workbook_path: Path,
    sheet_name: str | None,
    header_row: int,
    time_header: str,
    write_normalized_time: bool,
) -> tuple[object, object, int]:
    wb = load_workbook(workbook_path)
    ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]
    time_col_idx = find_time_column(ws, header_row, time_header)
    max_col = ws.max_column

    snapshots: list[RowSnapshot] = []
    errors: list[str] = []
    for row_idx in range(header_row + 1, ws.max_row + 1):
        row_values = [ws.cell(row_idx, col_idx).value for col_idx in range(1, max_col + 1)]
        if all(value is None for value in row_values):
            continue
        try:
            snapshots.append(build_row_snapshot(ws, row_idx, time_col_idx, max_col))
        except TimeParseError as exc:
            display = ws.cell(row_idx, time_col_idx).value
            errors.append(f"row {row_idx}: {display!r} -> {exc}")

    if errors:
        joined = "\n".join(errors)
        raise TimeParseError(f"Failed to parse {len(errors)} time cell(s):\n{joined}")

    snapshots.sort(key=lambda row: (row.time_span.sort_key, row.index))

    write_row_indices = list(range(header_row + 1, header_row + 1 + len(snapshots)))
    for offset, snapshot in zip(write_row_indices, snapshots):
        for col_idx, cell_snapshot in enumerate(snapshot.cells, start=1):
            target = ws.cell(offset, col_idx)
            target.value = cell_snapshot.value
            target._style = copy(cell_snapshot.style)
            target._hyperlink = copy(cell_snapshot.hyperlink) if cell_snapshot.hyperlink else None
            target.comment = copy(cell_snapshot.comment) if cell_snapshot.comment else None

        if write_normalized_time:
            time_cell = ws.cell(offset, time_col_idx)
            time_cell.value = snapshot.time_span.normalized
            time_cell.number_format = "@"

    for row_idx in range(header_row + 1 + len(snapshots), ws.max_row + 1):
        for col_idx in range(1, max_col + 1):
            target = ws.cell(row_idx, col_idx)
            target.value = None
            target._hyperlink = None
            target.comment = None

    return wb, ws, len(snapshots)


def resolve_output_path(input_path: Path, output_path: Path | None, in_place: bool) -> Path:
    if in_place and output_path is not None:
        raise ValueError("Use either --in-place or --output, not both.")
    if in_place:
        return input_path
    if output_path is not None:
        return output_path
    return input_path.with_name(f"{input_path.stem}.sorted{input_path.suffix}")


def parse_args(argv: Sequence[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Parse the Time column and sort Excel rows chronologically."
    )
    parser.add_argument(
        "excel_path",
        nargs="?",
        default=r"D:\Algo\Notes\society\history\中国.xlsx",
        help="Excel file to sort. Defaults to the history reference workbook.",
    )
    parser.add_argument(
        "--sheet",
        help="Sheet name to sort. Defaults to the first worksheet.",
    )
    parser.add_argument(
        "--header-row",
        type=int,
        default=1,
        help="Header row number. Defaults to 1.",
    )
    parser.add_argument(
        "--time-header",
        default=TIME_HEADER,
        help="Header text for the time column. Defaults to 'Time'.",
    )
    parser.add_argument(
        "--output",
        help="Write the sorted workbook to this path. Default is '<name>.sorted.xlsx'.",
    )
    parser.add_argument(
        "--in-place",
        action="store_true",
        help="Overwrite the input workbook in place.",
    )
    parser.add_argument(
        "--backup",
        action="store_true",
        help="Create a .bak copy before overwriting in place.",
    )
    parser.add_argument(
        "--write-normalized-time",
        action="store_true",
        help="Rewrite Time cells using canonical output like '-305.1.4' and '2001.1.5'.",
    )
    return parser.parse_args(argv)


def main(argv: Sequence[str] | None = None) -> int:
    args = parse_args(argv or sys.argv[1:])
    input_path = Path(args.excel_path)
    if not input_path.exists():
        print(f"Input workbook not found: {input_path}", file=sys.stderr)
        return 1

    output_path = resolve_output_path(
        input_path=input_path,
        output_path=Path(args.output) if args.output else None,
        in_place=args.in_place,
    )

    try:
        wb, ws, row_count = sort_sheet_rows(
            workbook_path=input_path,
            sheet_name=args.sheet,
            header_row=args.header_row,
            time_header=args.time_header,
            write_normalized_time=args.write_normalized_time,
        )
    except Exception as exc:  # pragma: no cover - CLI reporting path.
        print(str(exc), file=sys.stderr)
        return 1

    if args.in_place and args.backup:
        backup_path = input_path.with_suffix(input_path.suffix + ".bak")
        shutil.copy2(input_path, backup_path)
        print(f"Backup created: {backup_path}")

    wb.save(output_path)
    print(
        f"Sorted {row_count} data row(s) in sheet '{ws.title}' by '{args.time_header}'. "
        f"Saved to: {output_path}"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
