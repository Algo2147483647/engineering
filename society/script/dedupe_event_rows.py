#!/usr/bin/env python
"""Find overlapping event rows in Excel and keep the more informative one.

The script only uses the `Time` and `Event` columns to decide whether two rows
describe the same event:
- their time intervals overlap, and
- their event titles are identical after normalization, or one clearly contains
  the other.

When duplicates are found, the retained row is chosen by:
1. richer time precision,
2. more detailed event title,
3. more complete row content.
"""

from __future__ import annotations

import argparse
import dataclasses
import re
import sys
from collections import defaultdict
from copy import copy
from pathlib import Path
from typing import Sequence

from openpyxl import load_workbook

from sort_excel_by_time import TimeParseError, TimePoint, TimeSpan, parse_time_span


TIME_HEADER = "Time"
EVENT_HEADER = "Event"

EVENT_CLEAN_RE = re.compile(r"[\\s\"'“”‘’《》〈〉「」『』（）()【】\\[\\]{}，,。；;：:、·!！?？/\\\\|-]+")


@dataclasses.dataclass
class CellSnapshot:
    value: object
    style: object
    hyperlink: object
    comment: object


@dataclasses.dataclass
class RowInfo:
    row_index: int
    time_text: str
    event_text: str
    normalized_event: str
    time_span: TimeSpan
    interval_start: tuple[int | float, int, int]
    interval_end: tuple[int | float, int, int]
    time_precision: tuple[int, int]
    detail_score: tuple[int, int, int, int, int]
    cells: list[CellSnapshot]


class UnionFind:
    def __init__(self, values: list[int]) -> None:
        self.parent = {value: value for value in values}

    def find(self, value: int) -> int:
        root = value
        while self.parent[root] != root:
            root = self.parent[root]
        while self.parent[value] != value:
            parent = self.parent[value]
            self.parent[value] = root
            value = parent
        return root

    def union(self, a: int, b: int) -> None:
        ra = self.find(a)
        rb = self.find(b)
        if ra != rb:
            if ra < rb:
                self.parent[rb] = ra
            else:
                self.parent[ra] = rb


def normalize_event_text(text: object) -> str:
    cleaned = EVENT_CLEAN_RE.sub("", str(text or "").strip())
    return cleaned.lower()


def days_in_month(month: int) -> int:
    if month == 2:
        return 29
    if month in {4, 6, 9, 11}:
        return 30
    return 31


def point_floor(point: TimePoint) -> tuple[int | float, int, int]:
    month = point.month if point.month is not None else 1
    day = point.day if point.day is not None else 1
    return (point.year, month, day)


def point_ceil(point: TimePoint) -> tuple[int | float, int, int]:
    if point.month is None:
        return (point.year, 12, 31)
    if point.day is None:
        return (point.year, point.month, days_in_month(point.month))
    return (point.year, point.month, point.day)


def span_interval(span: TimeSpan) -> tuple[tuple[int | float, int, int], tuple[int | float, int, int]]:
    return point_floor(span.points[0]), point_ceil(span.points[-1])


def intervals_overlap(a: RowInfo, b: RowInfo) -> bool:
    return not (a.interval_end < b.interval_start or b.interval_end < a.interval_start)


def event_titles_overlap(a: str, b: str) -> bool:
    if not a or not b:
        return False
    if a == b:
        return True

    short, long = (a, b) if len(a) <= len(b) else (b, a)
    if len(short) < 4:
        return False
    if short in long and (len(short) / max(len(long), 1)) >= 0.45:
        return True
    return False


def event_titles_equal(a: str, b: str) -> bool:
    return bool(a) and a == b


def event_titles_contained(a: str, b: str) -> bool:
    if not a or not b or a == b:
        return False
    short, long = (a, b) if len(a) <= len(b) else (b, a)
    return len(short) >= 4 and short in long and (len(short) / max(len(long), 1)) >= 0.45


def same_start_family(a: RowInfo, b: RowInfo) -> bool:
    left = a.time_span.points[0]
    right = b.time_span.points[0]
    if left.year != right.year:
        return False
    if left.month is not None and right.month is not None and left.month != right.month:
        return False
    if left.day is not None and right.day is not None and left.day != right.day:
        return False
    return True


def point_precision(point: TimePoint) -> int:
    score = 0
    if point.month is not None:
        score += 2
    elif point.qualifier:
        score += 1
    if point.day is not None:
        score += 1
    if point.qualifier in {"前后", "以前", "以后"}:
        score -= 1
    return score


def span_precision(span: TimeSpan) -> tuple[int, int]:
    return (sum(point_precision(point) for point in span.points), len(span.points))


def row_content_score(values: list[object]) -> tuple[int, int]:
    nonempty = 0
    total_chars = 0
    for value in values:
        if value is None:
            continue
        text = str(value).strip()
        if not text:
            continue
        nonempty += 1
        total_chars += len(text)
    return nonempty, total_chars


def build_row_info(ws, row_index: int, time_col: int, event_col: int) -> RowInfo:
    values = [ws.cell(row_index, col).value for col in range(1, ws.max_column + 1)]
    time_text = str(values[time_col - 1] or "").strip()
    event_text = str(values[event_col - 1] or "").strip()
    if not time_text or not event_text:
        raise TimeParseError(f"Missing Time or Event at row {row_index}")

    time_span = parse_time_span(time_text)
    interval_start, interval_end = span_interval(time_span)
    normalized_event = normalize_event_text(event_text)
    nonempty, total_chars = row_content_score(values)

    cells = []
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row_index, col)
        cells.append(
            CellSnapshot(
                value=cell.value,
                style=copy(cell._style),
                hyperlink=copy(cell.hyperlink) if cell.hyperlink else None,
                comment=copy(cell.comment) if cell.comment else None,
            )
        )

    return RowInfo(
        row_index=row_index,
        time_text=time_text,
        event_text=event_text,
        normalized_event=normalized_event,
        time_span=time_span,
        interval_start=interval_start,
        interval_end=interval_end,
        time_precision=span_precision(time_span),
        detail_score=(
            span_precision(time_span)[0],
            span_precision(time_span)[1],
            len(normalized_event),
            nonempty,
            total_chars,
        ),
        cells=cells,
    )


def find_column(ws, header_name: str, header_row: int) -> int:
    for col in range(1, ws.max_column + 1):
        value = ws.cell(header_row, col).value
        if str(value).strip() == header_name:
            return col
    raise KeyError(f"Could not find header {header_name!r}")


def collect_rows(ws, header_row: int, time_col: int, event_col: int) -> tuple[list[RowInfo], list[str]]:
    rows: list[RowInfo] = []
    errors: list[str] = []
    for row_index in range(header_row + 1, ws.max_row + 1):
        values = [ws.cell(row_index, col).value for col in range(1, ws.max_column + 1)]
        if all(value is None or str(value).strip() == "" for value in values):
            continue
        try:
            rows.append(build_row_info(ws, row_index, time_col, event_col))
        except Exception as exc:
            errors.append(f"row {row_index}: {exc}")
    return rows, errors


def cluster_duplicates(rows: list[RowInfo]) -> dict[int, list[RowInfo]]:
    uf = UnionFind([row.row_index for row in rows])
    sorted_rows = sorted(rows, key=lambda row: (row.interval_start, row.interval_end, row.row_index))
    for i, left in enumerate(sorted_rows):
        for right in sorted_rows[i + 1 :]:
            if right.interval_start > left.interval_end:
                break
            if not intervals_overlap(left, right):
                continue
            if event_titles_equal(left.normalized_event, right.normalized_event):
                uf.union(left.row_index, right.row_index)
                continue
            if event_titles_contained(left.normalized_event, right.normalized_event) and same_start_family(left, right):
                uf.union(left.row_index, right.row_index)

    groups: dict[int, list[RowInfo]] = defaultdict(list)
    for row in rows:
        groups[uf.find(row.row_index)].append(row)
    return groups


def choose_best_row(group: list[RowInfo]) -> RowInfo:
    return max(group, key=lambda row: (row.detail_score, -row.row_index))


def rewrite_workbook(
    wb,
    ws,
    header_row: int,
    kept_rows: list[RowInfo],
    report_lines: list[str],
    report_sheet_name: str,
) -> None:
    max_col = ws.max_column

    for target_row, source_row in enumerate(kept_rows, start=header_row + 1):
        for col, snapshot in enumerate(source_row.cells, start=1):
            cell = ws.cell(target_row, col)
            cell.value = snapshot.value
            cell._style = copy(snapshot.style)
            cell._hyperlink = copy(snapshot.hyperlink) if snapshot.hyperlink else None
            cell.comment = copy(snapshot.comment) if snapshot.comment else None

    for row_index in range(header_row + 1 + len(kept_rows), ws.max_row + 1):
        for col in range(1, max_col + 1):
            cell = ws.cell(row_index, col)
            cell.value = None
            cell._hyperlink = None
            cell.comment = None

    if report_sheet_name in wb.sheetnames:
        del wb[report_sheet_name]
    report_ws = wb.create_sheet(report_sheet_name)
    report_ws.append(["Action", "Kept row", "Removed row", "Kept Time", "Removed Time", "Kept Event", "Removed Event"])
    for line in report_lines:
        report_ws.append(line.split("\t"))


def parse_args(argv: Sequence[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Deduplicate overlapping event rows in an Excel workbook.")
    parser.add_argument(
        "excel_path",
        nargs="?",
        default=r"D:\Algo\Notes\society\history\Event.xlsx",
        help="Workbook to process. Defaults to Event.xlsx.",
    )
    parser.add_argument(
        "--output",
        help="Output workbook path. Defaults to '<name>.deduped.xlsx'.",
    )
    parser.add_argument(
        "--sheet",
        help="Worksheet name. Defaults to the first worksheet.",
    )
    parser.add_argument(
        "--header-row",
        type=int,
        default=1,
        help="Header row number. Defaults to 1.",
    )
    parser.add_argument(
        "--time-header",
        default=TIME_HEADER,
        help="Header name for the time column.",
    )
    parser.add_argument(
        "--event-header",
        default=EVENT_HEADER,
        help="Header name for the event column.",
    )
    parser.add_argument(
        "--report-sheet",
        default="_dedupe_report",
        help="Sheet name used for the dedupe report.",
    )
    return parser.parse_args(argv)


def main(argv: Sequence[str] | None = None) -> int:
    args = parse_args(argv or sys.argv[1:])
    input_path = Path(args.excel_path)
    if not input_path.exists():
        print(f"Input workbook not found: {input_path}", file=sys.stderr)
        return 1

    output_path = Path(args.output) if args.output else input_path.with_name(f"{input_path.stem}.deduped{input_path.suffix}")

    wb = load_workbook(input_path)
    ws = wb[args.sheet] if args.sheet else wb[wb.sheetnames[0]]
    time_col = find_column(ws, args.time_header, args.header_row)
    event_col = find_column(ws, args.event_header, args.header_row)

    rows, errors = collect_rows(ws, args.header_row, time_col, event_col)
    if errors:
        print("\n".join(errors), file=sys.stderr)
        return 1

    groups = cluster_duplicates(rows)
    kept: list[RowInfo] = []
    removed_count = 0
    report_lines: list[str] = []

    for root, group in groups.items():
        if len(group) == 1:
            kept.append(group[0])
            continue

        winner = choose_best_row(group)
        kept.append(winner)
        for row in sorted(group, key=lambda item: item.row_index):
            if row.row_index == winner.row_index:
                continue
            removed_count += 1
            report_lines.append(
                "\t".join(
                    [
                        "Removed duplicate",
                        str(winner.row_index),
                        str(row.row_index),
                        winner.time_text,
                        row.time_text,
                        winner.event_text,
                        row.event_text,
                    ]
                )
            )

    kept.sort(key=lambda row: row.row_index)
    rewrite_workbook(
        wb=wb,
        ws=ws,
        header_row=args.header_row,
        kept_rows=kept,
        report_lines=report_lines,
        report_sheet_name=args.report_sheet,
    )
    wb.save(output_path)

    duplicate_groups = sum(1 for group in groups.values() if len(group) > 1)
    print(
        f"Processed {len(rows)} data row(s); found {duplicate_groups} duplicate group(s); "
        f"removed {removed_count} row(s). Saved to: {output_path}"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
